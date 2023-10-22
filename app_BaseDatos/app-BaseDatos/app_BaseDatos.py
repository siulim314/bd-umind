from openpyxl import Workbook
from openpyxl.styles import Font

# =========================================
# CONFIGURACION
# =========================================
wb = Workbook()
ws = wb.active


# =========================================
# INPUTS
# =========================================
cabecera = ["PRUEBA1","PRUEBA2","PRUEBA3"]
path = "C:\\Users\\luism\\Desktop\\Umind\\app\\app_BaseDatos\\archivos\\BaseDatos_Umind.xlsx"


# =========================================
# FUNCIONES
# =========================================
def insertar_cabecera(cabecera):
    cont = 0
    for cab in cabecera:
        cont = cont+1
        # Insertar nuevo valor de cabecera
        celda = ws.cell(row=1, column=cont)
        celda.value = cab
        conf_fuente_columnas(celda)

def insertar_row(row,col,data):

    celda = ws.cell(row=row, column=col)
    celda.value = data
    conf_fuente_columnas(celda)

def eliminar_row(celda):

    ws.delete_rows(celda.row)

def conf_fuente_filas(celda):

    font = Font(name='Arial', size=10, color='00000000', bold=True)
    celda.font = font

def insertar_col(cabecera,data):
    # Insertar nueva columna
    new_col = len(cabecera)+1
    ws.insert_cols(new_col)
    # Insertar nuevo valor de cabecera
    celda = ws.cell(row=1, column=new_col)
    celda.value=data
    conf_fuente_columnas(celda)

def eliminar_col(celda):

    ws.delete_cols(celda.col)

def conf_fuente_columnas(celda):

    font = Font(name='Arial', size=10, color='00000000', bold=True)
    celda.font = font

def guardar_file(path):

    # Save the file
    wb.save(path)



def main():
    insertar_cabecera(cabecera)
    insertar_row(2,1,"FILA")
    insertar_col(cabecera,"COLUMNA")
    guardar_file(path)
main()
