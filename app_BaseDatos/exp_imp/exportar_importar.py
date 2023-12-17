import sqlite3
from model.conexion_db import ConexionDB
import pandas as pd
from mensajes import mensaje
from tkinter import filedialog
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import constantes.constantes as cte

ruta_plantilla = cte.ruta_plantilla
name_archivo = cte.ruta_name_archivo

# Crear un nuevo archivo de Excel
worbook = openpyxl.Workbook()
sheet = worbook.active


def exportar(flag,obtener_parametros,obtener_contenido,obtener_tipos_trastornos,obtener_categoria):
    # flag = 1 -> exportar con contenido
    # flag = 0 -> exportar sin contenido (plantilla)

    # Obtiene los datos de la base de datos
    column_names = obtener_parametros
    contenido = obtener_contenido
    column_tipotras = obtener_tipos_trastornos
    print(f"Obtener columnas {column_tipotras}")

    contenido = sorted(contenido)

    # Rellenamos el excel columna a columna
    try:
        for c in range(1,len(column_names)):
            if c==1:
                column_names[1] = "categoria-trastornos"
            else:
                pass

            # Escribir los parámetros
            cab = sheet.cell(row=1, column=c)
            cab.value = column_names[c]

        for r in range(len(contenido)):
            for c in range(len(column_names)):
                c_1 = c+1
                r_2 = r+2
                if flag == True:
                    # Escribir el contenido
                    cont = sheet.cell(row=r_2, column=c_1)
                    cont.value = contenido[r][c_1]

                if flag == False:
                    # Función para la plantilla
                    cont = sheet.cell(row=r_2, column=c_1)
                    cont.value = ""

        # Crear una regla de validación de datos para la celda B2
        validation_tipotrans = DataValidation(
            type="list",
            formula1=f'"{",".join(column_tipotras)}"',
            allow_blank=True
        )

        validation_tipotrans.add('A2:A1000000')
        sheet.add_data_validation(validation_tipotrans)


        if flag == True:
            worbook.save(name_archivo)
        elif flag == False:
            worbook.save(ruta_plantilla + name_archivo)
    except:
        mensaje('Exportar archivo',
                'No ha sido posible exportar',
                2)
    else:
        mensaje('Exportar archivo',
                'Archivo exportado',
                0)

def importar(nombre_tabla_trastornos):

    conexiones = ConexionDB()

    # Conectarse a la base de datos SQLite (creará la base de datos si no existe)
    conexion = sqlite3.connect(conexiones.base_datos)
    # El usuario eligira el excel a importar
    archivo = filedialog.askopenfilename(filetypes=[("Archivos de texto", "*.xlsx"),
                                                    ("Todos los archivos", "*.*")])

    try:
        # Leer el archivo de Excel en un DataFrame
        df = pd.read_excel(archivo)

        # Guardar el DataFrame en la base de datos SQLite
        df.to_sql(nombre_tabla_trastornos, conexion, if_exists='append', index=False)

    except:
        mensaje('Importar archivo',
                'No ha sido posible importar',
                2)
    else:
        mensaje('Importar archivo',
                f'Archivo: {name_archivo} importado',
                0)

    # Cerrar la conexión a la base de datos
    conexion.close()
    conexiones.cerrar()